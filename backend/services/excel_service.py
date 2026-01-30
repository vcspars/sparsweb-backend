from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import os
from pathlib import Path
from sqlalchemy.orm import Session
from database import (
    NewsletterSubscription,
    ContactForm,
    BrochureForm,
    ProductProfileForm,
    TalkToSalesForm
)

class ExcelService:
    def __init__(self):
        self.excel_dir = Path(__file__).parent.parent / "exports"
        self.excel_dir.mkdir(exist_ok=True)
        self.filename = "SPARS_Excel_DB.xlsx"

    def export_all_forms(self, db: Session) -> str:
        """Export all forms to Excel with separate sheets - overwrites existing file"""
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Export each form type to separate sheet (6 forms total)
        self._export_newsletter(wb, db)
        self._export_contact_forms(wb, db)
        self._export_demo_requests(wb, db)
        self._export_brochure_forms(wb, db)
        self._export_product_profile_forms(wb, db)
        self._export_talk_to_sales_forms(wb, db)
        
        # Save file (overwrites if exists)
        file_path = self.excel_dir / self.filename
        wb.save(file_path)
        return str(file_path)

    def _export_newsletter(self, wb: Workbook, db: Session):
        """Export newsletter subscriptions"""
        ws = wb.create_sheet("Newsletter Subscriptions")
        subscriptions = db.query(NewsletterSubscription).order_by(NewsletterSubscription.subscribed_at.desc()).all()
        
        # Headers
        headers = ["ID", "Email", "Subscribed At"]
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Data
        for sub in subscriptions:
            ws.append([
                sub.id,
                sub.email,
                sub.subscribed_at.strftime("%Y-%m-%d %H:%M:%S") if sub.subscribed_at else ""
            ])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _export_contact_forms(self, wb: Workbook, db: Session):
        """Export contact forms (excluding demo requests)"""
        ws = wb.create_sheet("Contact Forms")
        # Get contact forms where demo_date is None (not demo requests)
        forms = db.query(ContactForm).filter(ContactForm.demo_date == None).order_by(ContactForm.submitted_at.desc()).all()
        
        headers = ["ID", "First Name", "Last Name", "Email", "Phone", "Company", "Company Size", "Message", "Submitted At"]
        ws.append(headers)
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        for form in forms:
            ws.append([
                form.id,
                form.first_name or "",
                form.last_name or "",
                form.email,
                form.phone or "",
                form.company or "",
                form.company_size or "",
                form.message or "",
                form.submitted_at.strftime("%Y-%m-%d %H:%M:%S") if form.submitted_at else ""
            ])
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _export_demo_requests(self, wb: Workbook, db: Session):
        """Export demo request forms"""
        ws = wb.create_sheet("Demo Requests")
        # Get contact forms where demo_date is not None (demo requests)
        forms = db.query(ContactForm).filter(ContactForm.demo_date != None).order_by(ContactForm.submitted_at.desc()).all()
        
        headers = ["ID", "First Name", "Last Name", "Email", "Phone", "Company", "Company Size", "Preferred Demo Date", "Additional Information", "Current ERP System", "Number of Warehouses", "Expected Number of Users", "Specific Requirements or Challenges", "Implementation Timeline", "Submitted At"]
        ws.append(headers)
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        for form in forms:
            ws.append([
                form.id,
                form.first_name or "",
                form.last_name or "",
                form.email,
                form.phone or "",
                form.company or "",
                form.company_size or "",
                form.demo_date or "",
                form.message or "",
                form.current_system or "",
                form.warehouses or "",
                form.users or "",
                form.requirements or "",
                form.timeline or "",
                form.submitted_at.strftime("%Y-%m-%d %H:%M:%S") if form.submitted_at else ""
            ])
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _export_brochure_forms(self, wb: Workbook, db: Session):
        """Export brochure forms"""
        ws = wb.create_sheet("Brochure Requests")
        forms = db.query(BrochureForm).order_by(BrochureForm.submitted_at.desc()).all()
        
        headers = ["ID", "First Name", "Last Name", "Email", "Company", "Phone", "Job Role", "Agreed to Marketing", "Submitted At"]
        ws.append(headers)
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        for form in forms:
            ws.append([
                form.id,
                form.first_name or "",
                form.last_name or "",
                form.email,
                form.company,
                form.phone or "",
                form.job_role or "",
                "Yes" if form.agreed_to_marketing else "No",
                form.submitted_at.strftime("%Y-%m-%d %H:%M:%S") if form.submitted_at else ""
            ])
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _export_product_profile_forms(self, wb: Workbook, db: Session):
        """Export product profile forms"""
        ws = wb.create_sheet("Product Profile Requests")
        forms = db.query(ProductProfileForm).order_by(ProductProfileForm.submitted_at.desc()).all()
        
        headers = [
            "ID", "First Name", "Last Name", "Email", "Phone", "Job Title",
            "Company Name", "Industry", "Company Size", "Website", "Address",
            "Current System", "Warehouses", "Users", "Requirements", "Timeline", "Submitted At"
        ]
        ws.append(headers)
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        for form in forms:
            ws.append([
                form.id,
                form.first_name,
                form.last_name,
                form.email,
                form.phone,
                form.job_title or "",
                form.company_name,
                form.industry or "",
                form.company_size or "",
                form.website or "",
                form.address or "",
                form.current_system or "",
                form.warehouses or "",
                form.users or "",
                form.requirements or "",
                form.timeline or "",
                form.submitted_at.strftime("%Y-%m-%d %H:%M:%S") if form.submitted_at else ""
            ])
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _export_talk_to_sales_forms(self, wb: Workbook, db: Session):
        """Export talk to sales forms"""
        ws = wb.create_sheet("Talk to Sales")
        forms = db.query(TalkToSalesForm).order_by(TalkToSalesForm.submitted_at.desc()).all()
        
        headers = [
            "ID", "First Name", "Last Name", "Email", "Phone", "Company", "Company Size", "Additional Information",
            "Current ERP System", "Number of Warehouses", "Expected Number of Users",
            "Specific Requirements or Challenges", "Implementation Timeline", "Submitted At"
        ]
        ws.append(headers)
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        for form in forms:
            ws.append([
                form.id,
                form.first_name or "",
                form.last_name or "",
                form.email,
                form.phone,
                form.company or "",
                form.company_size or "",
                form.additional_information or "",
                form.current_system or "",
                form.warehouses or "",
                form.users or "",
                form.requirements or "",
                form.timeline or "",
                form.submitted_at.strftime("%Y-%m-%d %H:%M:%S") if form.submitted_at else ""
            ])
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

